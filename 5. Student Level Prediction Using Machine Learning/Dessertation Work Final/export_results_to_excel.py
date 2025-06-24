
"""
This script gathers all relevant results (metrics, confusion matrices, SHAP, LIME, etc.)
and exports them into a single Excel file with multiple sheets + embedded plots.
"""

import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils.dataframe import dataframe_to_rows

# Create EXCEL directory if not exists
os.makedirs("EXCEL", exist_ok=True)
excel_path = "EXCEL/The_Student_Dataset_Results.xlsx"

# Prepare ExcelWriter
with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    # 1. Write overall metrics
    if 'metrics_df' in globals():
        metrics_df.to_excel(writer, sheet_name="Overall_Metrics")

    # 2. Confusion Matrices as TP, FP, FN, TN table
    if 'confusion_df' in globals():
        confusion_df.to_excel(writer, sheet_name="Confusion_Matrices", index=False)

    # 3. Feature imbalance
    if 'imbalance_df' in globals():
        imbalance_df.to_excel(writer, sheet_name="Feature_Imbalance")

    # 4. SHAP Global Importance
    if 'importance_df' in globals():
        importance_df.to_excel(writer, sheet_name="SHAP_Global_Importance", index=False)

    # 5. SHAP Local
    if 'shap_values_lgbm' in globals():
        try:
            shap_local = pd.DataFrame(shap_values_lgbm[index_to_check].values, columns=["SHAP_Value"])
            shap_local["Feature"] = X_test.columns
            shap_local = shap_local[["Feature", "SHAP_Value"]]
            shap_local.to_excel(writer, sheet_name="SHAP_Local_Explanation", index=False)
        except:
            pass

    # 6. LIME Explanations
    if 'feature_importance_df' in globals():
        feature_importance_df.to_excel(writer, sheet_name="LIME_Explanations", index=False)

    # 7. Native Importance
    if 'native_fi_df' in globals():
        native_fi_df.to_excel(writer, sheet_name="Native_Importance", index=False)

    # 8. Permutation Importance
    for model_name in ['LightGBM', 'XGBoost']:
        csv_path = f'The_Student_Dataset_Permutation_Importance_{model_name}.csv'
        if os.path.exists(csv_path):
            perm_df = pd.read_csv(csv_path)
            perm_df.to_excel(writer, sheet_name=f'Permutation_Importance_{model_name}', index=False)

    # 9. PDP data
    for model_name in ['LightGBM', 'XGBoost']:
        pdp_path = f'The_Student_Dataset_PDP_{model_name}.png'
        if os.path.exists(pdp_path):
            sheet = writer.book.create_sheet(f'PDP_Features_{model_name}')
            sheet.cell(1, 1, f"PDP plot saved as image for {model_name}")

    # 10. ALE
    if os.path.exists("The_Student_Dataset_ALE_LightGBM_XGBoost.png"):
        writer.book.create_sheet("ALE_Data")
        writer.book["ALE_Data"].cell(1, 1, "ALE plot saved as image.")

    # 11. SHAP Interaction
    for model_name in ['LightGBM', 'XGBoost']:
        img_path = f'The_Student_Dataset_SHAP_Interaction_{model_name}.png'
        if os.path.exists(img_path):
            sheet = writer.book.create_sheet(f'SHAP_Interaction_{model_name}')
            sheet.cell(1, 1, f"SHAP interaction plot saved for {model_name}.")

    # 12. Surrogate model
    from pathlib import Path
    surrogate_accuracies = []
    for model_name in ['LightGBM', 'XGBoost']:
        surrogate_img = f'The_Student_Dataset_Global_Surrogate_{model_name}.png'
        if Path(surrogate_img).exists():
            surrogate_accuracies.append({"Model": model_name, "Surrogate Accuracy": "See Plot"})
    if surrogate_accuracies:
        surrogate_df = pd.DataFrame(surrogate_accuracies)
        surrogate_df.to_excel(writer, sheet_name="Surrogate_Accuracies", index=False)

    # 13. ROC curve data
    roc_data = []
    if 'roc_curves' in globals():
        for model, (fpr, tpr) in roc_curves.items():
            for f, t in zip(fpr, tpr):
                roc_data.append({"Model": model, "FPR": f, "TPR": t})
    if roc_data:
        pd.DataFrame(roc_data).to_excel(writer, sheet_name="ROC_Curve_Data", index=False)

    # 14. Predictions
    if 'data' in globals():
        data.to_excel(writer, sheet_name="Predictions", index=False)

    # 15. Correlation with Class
    if 'correlation_table' in globals():
        correlation_table.to_excel(writer, sheet_name="Correlation_with_Class", index=False)

    # 16. Text-based Confusion Matrix with Labels
    matrix_text = []
    if 'best_confusion_matrices' in globals():
        for model, cm in best_confusion_matrices.items():
            TN, FP, FN, TP = cm.ravel()
            matrix_text.append({
                "Model": model,
                "True Negative": TN,
                "False Positive": FP,
                "False Negative": FN,
                "True Positive": TP
            })
    if matrix_text:
        pd.DataFrame(matrix_text).to_excel(writer, sheet_name="Confusion_Matrix_Text", index=False)

print(f"All results exported to Excel: {excel_path}")
